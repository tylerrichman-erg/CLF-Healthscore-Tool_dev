import logging
from datetime import datetime, tzinfo, timedelta
from django.conf import settings
import pandas as pd
import numpy as np
import statistics
import io
import numbers
import os
from pandas import DataFrame
from django.db.models import Avg
from healthscore.data.ACSDataLoader import ACSDataLoader
from healthscore.data.CDCDataLoader import CDCDataLoader
from healthscore.data.EJScreenDataLoader import EJScreenDataLoader
from healthscore.models import Dataset, Metric, State, BRFSS, EducationMA, EducationMASubgroup, EducationCT,\
    EducationCTSubgroup, EducationRI, MetricValence, LifeExpectancy, Latch, SmartLocation, NHTS, ChildMortality,\
    PersonalHealthCare, ChildHealth, ERVisits, MaternityCare, MentalHealthCare, HouseholdBurden, Homelessness
from django.core.exceptions import ObjectDoesNotExist
from healthscore.services.data_service import DataService
from openpyxl import load_workbook


class ExcelService:

    """
    This class produces an Excel spreadsheet summary of various metrics, broken down by census tract, aggregated
    across tracts and state, and scored according to various methods.
    """
    def __init__(self, user):
        self.user = user
        self.log = logging.getLogger('requested')

    def has_access(self, user):
        return user.groups.filter(name='Admin').exists()

    def is_NEF(self, scenario):
        return scenario == 'NEF'

    def get_column_letter(self, column_int):
        start_index = 1   #  it can start either at 0 or at 1
        letter = ''
        while column_int > 25 + start_index:
            letter += chr(65 + int((column_int-start_index)/26) - 1)
            column_int = column_int - (int((column_int-start_index)/26))*26
        letter += chr(65 - start_index + (int(column_int)))
        return letter

    def set_formulae(self, wb, community, num_tracts, housing_data, school_district, scenario):
        hs_sheet = wb['HealthScore 2.0_' + community]

        # Identify the column in the scorecard that has the "% Points scored" value. This depends
        # on how many tracts were included in the analysis, as follows:
        # There are 4 columns per tract, plus 'all tracts' and state (and additionally county and national if this is
        # 'NEF' scenario...), plus the first column (metric name), and then points scored is the last column:
        # (4 * (num_tracts+2)) + 7 ==> 4*num_tracts + 15 (non-NEF)
        # (4 * (num_tracts+4)) + 7 ==> 4*num_tracts + 23 (NEF)
        col = 4*num_tracts + 23 if self.is_NEF(scenario) else 4*num_tracts + 15
        col_letter = self.get_column_letter(col)

        housing_row = len(housing_data) + 27
        housing_col = num_tracts*4 + 3
        housing_col_letter = self.get_column_letter(housing_col)
        valid_housing_data = housing_data is not None and len(housing_data) > 0 and housing_data[0]['unit_type'] != ''

        # These are the same for both advantaged and disadvantaged...
        hs_sheet.cell(15, 3).value = f"=IF('Raw data'!{col_letter}4=\"\", \"\", PRODUCT('Raw data'!{col_letter}4, D15))"
        hs_sheet.cell(16, 3).value = f"=IF('Raw data'!{col_letter}5=\"\", \"\", PRODUCT('Raw data'!{col_letter}5, D16))"
        hs_sheet.cell(17, 3).value = f"=IF('Raw data'!{col_letter}6=\"\", \"\", PRODUCT('Raw data'!{col_letter}6, D17))"
        hs_sheet.cell(18, 3).value = f"=IF('Raw data'!{col_letter}7=\"\", \"\", PRODUCT('Raw data'!{col_letter}7, D18))"
        hs_sheet.cell(19, 3).value = f"=IF('Raw data'!{col_letter}8=\"\", \"\", PRODUCT('Raw data'!{col_letter}8, D19))"
        hs_sheet.cell(20, 3).value = f"=IF('Raw data'!{col_letter}9=\"\", \"\", PRODUCT('Raw data'!{col_letter}9, D20))"
        hs_sheet.cell(21, 3).value = f"=IF('Raw data'!{col_letter}10=\"\", \"\", PRODUCT('Raw data'!{col_letter}10, D21))"
        hs_sheet.cell(22, 3).value = f"=IF('Raw data'!{col_letter}11=\"\", \"\", PRODUCT('Raw data'!{col_letter}11, D22))"
        hs_sheet.cell(23, 3).value = f"=IF('Raw data'!{col_letter}12=\"\", \"\", PRODUCT('Raw data'!{col_letter}12, D23))"
        hs_sheet.cell(24, 3).value = f"=IF('Raw data'!{col_letter}13=\"\", \"\", PRODUCT('Raw data'!{col_letter}13, D24))"
        hs_sheet.cell(25, 3).value = f"=IF('Raw data'!{col_letter}14=\"\", \"\", PRODUCT('Raw data'!{col_letter}14, D25))"
        hs_sheet.cell(26, 3).value = f"=IF('Raw data'!{col_letter}15=\"\", \"\", PRODUCT('Raw data'!{col_letter}15, D26))"
        hs_sheet.cell(30, 3).value = f"=IF('Raw data'!{col_letter}24=\"\", \"\", PRODUCT('Raw data'!{col_letter}24, D30))"
        hs_sheet.cell(31, 3).value = f"=PRODUCT(AVERAGE('Raw data'!{col_letter}16, 'Raw data'!{col_letter}17, 'Raw data'!{col_letter}18), D31)"
        hs_sheet.cell(32, 3).value = f"=IF('Raw data'!{col_letter}19=\"\", \"\", PRODUCT('Raw data'!{col_letter}19, D32))"

        if community == 'Advantaged':
            hs_sheet.cell(35, 3).value = f'''=IF('Raw data'!{col_letter}34 > 66.7, "Pass", "Fail")''' if school_district else "Data unavailable"
            hs_sheet.cell(36, 3).value = f"=PRODUCT('Raw data'!{col_letter}35, D36)" if school_district else "Data unavailable"

            hs_sheet.cell(38, 3).value = f"=IF('Raw data'!{col_letter}25=\"\", \"\", PRODUCT('Raw data'!{col_letter}25, D38))"
            hs_sheet.cell(39, 3).value = f"=IF('Raw data'!{col_letter}26=\"\", \"\", PRODUCT('Raw data'!{col_letter}26, D39))"
            hs_sheet.cell(40, 3).value = f"=IF('Raw data'!{col_letter}27=\"\", \"\", PRODUCT('Raw data'!{col_letter}27, D40))"
            hs_sheet.cell(41, 3).value = f"=IF('Raw data'!{col_letter}28=\"\", \"\", PRODUCT('Raw data'!{col_letter}28, D41))"
            hs_sheet.cell(42, 3).value = f"=IF('Raw data'!{col_letter}31=\"\", \"\", PRODUCT('Raw data'!{col_letter}31, D42))"
            hs_sheet.cell(43, 3).value = f"=IF('Raw data'!{col_letter}32=\"\", \"\", PRODUCT('Raw data'!{col_letter}32, D43))"
            hs_sheet.cell(44, 3).value = f"=IF('Raw data'!{col_letter}33=\"\", \"\", PRODUCT('Raw data'!{col_letter}33, D44))"
            hs_sheet.cell(71, 3).value = f"=PRODUCT('Housing Affordability'!{housing_col_letter}{housing_row}, D71)" if valid_housing_data else "Data unavailable"
            hs_sheet.cell(139, 3).value = "Data Unavailable"
        elif community == 'Disadvantaged':
            hs_sheet.cell(35, 3).value = f"=IF('Raw data'!{col_letter}25=\"\", \"\", PRODUCT('Raw data'!{col_letter}25, D35))"
            hs_sheet.cell(36, 3).value = f"=IF('Raw data'!{col_letter}26=\"\", \"\", PRODUCT('Raw data'!{col_letter}26, D36))"
            hs_sheet.cell(37, 3).value = f"=IF('Raw data'!{col_letter}27=\"\", \"\", PRODUCT('Raw data'!{col_letter}27, D37))"
            hs_sheet.cell(38, 3).value = f"=IF('Raw data'!{col_letter}28=\"\", \"\", PRODUCT('Raw data'!{col_letter}28, D38))"
            hs_sheet.cell(39, 3).value = f"=IF('Raw data'!{col_letter}29=\"\", \"\", PRODUCT('Raw data'!{col_letter}29, D39))"
            hs_sheet.cell(40, 3).value = f"=IF('Raw data'!{col_letter}30=\"\", \"\", PRODUCT('Raw data'!{col_letter}30, D40))"
            hs_sheet.cell(41, 3).value = f"=IF('Raw data'!{col_letter}31=\"\", \"\", PRODUCT('Raw data'!{col_letter}31, D41))"
            hs_sheet.cell(42, 3).value = f"=IF('Raw data'!{col_letter}32=\"\", \"\", PRODUCT('Raw data'!{col_letter}32, D42))"
            hs_sheet.cell(43, 3).value = f"=IF('Raw data'!{col_letter}33=\"\", \"\", PRODUCT('Raw data'!{col_letter}33, D43))"
            hs_sheet.cell(70, 3).value = f"=PRODUCT('Housing Affordability'!{housing_col_letter}{housing_row}, D70)" if valid_housing_data else "Data unavailable"
            hs_sheet.cell(138, 3).value = "Data Unavailable"

    def to_excel(self, user, address: str, state_fips: str, state_short_code: str, county: str, county_name: str, tracts: list,
                 primary_tract: str, block: str, school_district: str, community: str, housing_data: list,
                 buffer_radius: int, population_threshold: float, scenario: str, include_raw: bool):

        output = io.BytesIO()

        # If user has sufficient access (i.e. belongs to the right group), load the template spreadsheet,
        # and write out its entire contents to the output bytes
        writer = None
        if self.has_access(user):
            path_to_workbook = os.path.join(settings.BASE_DIR, 'healthscore.xlsx')
            wb = load_workbook(path_to_workbook)

            # Remove sheets that are not relevant. This is either advantaged or disadvantaged scenario.
            for sheet in wb.sheetnames:
                if (community == 'Advantaged' and 'Disadvantaged' in sheet) or\
                        (community == 'Disadvantaged' and 'Advantaged' in sheet):
                    ws = wb[sheet]
                    wb.remove(ws)

            self.set_formulae(wb, community, len(tracts), housing_data, school_district, scenario)

            wb.save(output)
            writer = pd.ExcelWriter(output, mode="a", engine="openpyxl", if_sheet_exists="overlay")
        else:
            writer = pd.ExcelWriter(output, mode="w", engine="openpyxl")

        df = self.construct_dataframe(state=state_short_code, tracts=tracts, county_name=county_name, scenario=scenario)

        vintages = DataService.vintages()

        # Find the state
        try:
            state = State.objects.get(short_code=state_short_code)
        except ObjectDoesNotExist:
            self.log.error("Couldn't find state with abbreviation " + state_short_code,
                           extra={'user': self.user.username})
            return None

        for t in tracts:

            # Add values that came from APIs etc.
            self.add_acs_all(vintage=vintages['ACS'], state=state, county=county, tract=t, base_df=df)
            self.add_cdc_tract(vintage=vintages['CDC'], state_fips=state.fips_code, county=county, tract=t, base_df=df)
            self.add_ejscreen_tract(vintage=vintages['EJScreen'], state_fips=state.fips_code, county=county, tract=t,
                                    base_df=df)
            self.add_life_exp_tract(vintage=vintages['LifeExpectancy'], state_fips=state.fips_code, county=county,
                                    tract=t, base_df=df)
            self.add_latch_tract(vintage=vintages['Latch'], state_fips=state.fips_code, county=county, tract=t,
                                 base_df=df)

        # All tracts, transit frequency EPA Smart Location
        self.add_smart_location_all(vintage=vintages['SmartLocation'], state_fips=state.fips_code, county=county,
                                    primary=primary_tract, block=block, base_df=df)

        if self.is_NEF(scenario):
            # County, state, national level child mortality
            county_fips = state_fips + county
            cm_county = ChildMortality.objects.filter(dataset__vintage=vintages['ChildMortality'],
                                                      county_fips=county_fips).first()
            df.loc['Child Mortality', (county_name, 'EST')] = cm_county.value
            df.loc['Child Mortality', (county_name, 'MOE')] = cm_county.moe

            cm_state = ChildMortality.objects.filter(dataset__vintage=vintages['ChildMortality'],
                                                     state_fips=state_fips, county_fips=None).first()
            df.loc['Child Mortality', (state_short_code, 'EST')] = cm_state.value
            df.loc['Child Mortality', (state_short_code, 'MOE')] = cm_state.moe

            cm_us = ChildMortality.objects.filter(dataset__vintage=vintages['ChildMortality'],
                                                  state_fips=None, county_fips=None).first()
            df.loc['Child Mortality', ('US', 'EST')] = cm_us.value
            df.loc['Child Mortality', ('US', 'MOE')] = cm_us.moe
            df.loc['Child Mortality', 'Source'] = 'County Health'

            # State personal healthcare cost per capita. In order to score this metric, we have to
            # calculate the ratio of (per capita cost) / (median household income) for each state,
            # and then place them in quintiles, assigning a percentage of points based on which
            # bucket this state lands in. Note that the state-level Median Household Income used is a
            # single value, unlike the range+band that we place in the state column for that metric.
            # NOTE: the same methodology will be applied to childhood obesity and asthma.
            acs_incomes = list(Metric.objects.filter(dataset__vintage=vintages['ACS'],
                                                     dataset__data_source__name='ACS Detail',
                                                     name__exact='Median Household Income',
                                                     geoid__regex='^\d\d$'))

            incomes_by_geoid = {a.geoid: a.value for a in acs_incomes}

            phc_states = list(PersonalHealthCare.objects.filter(dataset__vintage=vintages['PersonalHealthCare'],
                                                                state_fips__isnull=False).all())
            per_capita_data = []
            for phc_state in phc_states:
                if phc_state.state_fips in incomes_by_geoid:
                    income = incomes_by_geoid[phc_state.state_fips]
                    per_capita_data.append([phc_state.state_fips, phc_state.value / income])

            quintile_df = pd.DataFrame(columns=['geoid', 'ratio'], data=per_capita_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['ratio'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            df.loc['Total healthcare cost per capita per household income', (state_short_code, 'EST')] = \
                row.iloc[0].ratio
            df.loc['Total healthcare cost per capita per household income', '% Points'] = \
                1 - (row.iloc[0].quintile * 0.25)
            df.loc['Total healthcare cost per capita per household income', 'Source'] = 'CMS'

            # Childhood obesity and asthma
            nsch_states = list(ChildHealth.objects.filter(dataset__vintage=vintages['NSCH']).all())
            obesity_data = []
            for obesity_state in nsch_states:
                if obesity_state.state_fips in incomes_by_geoid:
                    obesity_data.append([obesity_state.state_fips, obesity_state.obesity_value])

            quintile_df = pd.DataFrame(columns=['geoid', 'ratio'], data=obesity_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['ratio'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            df.loc['Child obesity', (state_short_code, 'EST')] = row.iloc[0].ratio
            df.loc['Child obesity', '% Points'] = row.iloc[0].quintile * 0.25
            df.loc['Child obesity', 'Source'] = 'NSCH'

            asthma_data = []
            for asthma_state in nsch_states:
                if asthma_state.state_fips in incomes_by_geoid:
                    asthma_data.append([asthma_state.state_fips, asthma_state.asthma_value])

            quintile_df = pd.DataFrame(columns=['geoid', 'ratio'], data=asthma_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['ratio'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            df.loc['Child asthma', (state_short_code, 'EST')] = row.iloc[0].ratio
            df.loc['Child asthma', '% Points'] = row.iloc[0].quintile * 0.25
            df.loc['Child asthma', 'Source'] = 'NSCH'

            # Emergency room visits
            kff_states = list(ERVisits.objects.filter(dataset__vintage=vintages['ERVisits'],
                                                      state_fips__isnull=False).all())
            er_data = []
            for er_state in kff_states:
                state_value = er_state.state_local_value if er_state.state_local_value is not None else 0
                profit_value = er_state.for_profit_value if er_state.for_profit_value is not None else 0
                non_profit_value = er_state.non_profit_value if er_state.non_profit_value is not None else 0
                total = state_value + profit_value + non_profit_value
                er_data.append([er_state.state_fips, total])

            quintile_df = pd.DataFrame(columns=['geoid', 'total'], data=er_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['total'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            df.loc['ER Visits', (state_short_code, 'EST')] = row.iloc[0].total
            df.loc['ER Visits', '% Points'] = 1 - (row.iloc[0].quintile * 0.25)

            er_us = ERVisits.objects.filter(dataset__vintage=vintages['ERVisits'],
                                                  state_fips=None).first()
            df.loc['ER Visits', ('US', 'EST')] = er_us.state_local_value + \
                                                 er_us.non_profit_value + \
                                                 er_us.for_profit_value

            df.loc['ER Visits', 'Source'] = 'KFF'

            # Maternity Care
            mc_county = MaternityCare.objects.filter(dataset__vintage=vintages['MaternityCare'],
                                               county_fips=county_fips).first()
            df.loc['Maternity Care', (county_name, 'EST')] = mc_county.care_level
            if mc_county.care_level == "Maternity Care Desert":
                df.loc['Maternity Care', '% Points'] = 0
            elif mc_county.care_level == "Low Access to Care":
                df.loc['Maternity Care', '% Points'] = 1/3
            elif mc_county.care_level == "Moderate Access to Care":
                df.loc['Maternity Care', '% Points'] = 2/3
            else:
                df.loc['Maternity Care', '% Points'] = 1
            df.loc['Maternity Care', 'Source'] = 'MoD'

            # Access to Mental Health Care
            mha_states = list(MentalHealthCare.objects.filter(dataset__vintage=vintages['MentalHealthCare']).all())
            mha_data = []
            for mha_state in mha_states:
                mha_data.append([mha_state.state_fips, mha_state.rank])

            quintile_df = pd.DataFrame(columns=['geoid', 'state_rank'], data=mha_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['state_rank'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            df.loc['Access to Mental Health Care', (state_short_code, 'EST')] = row.iloc[0].state_rank
            df.loc['Access to Mental Health Care', '% Points'] = 1 - (row.iloc[0].quintile * 0.25)
            df.loc['Access to Mental Health Care', 'Source'] = 'MHA'

            # Household Burden
            for tract in tracts:
                geoid = county_fips + tract
                burden = HouseholdBurden.objects.filter(dataset__vintage=vintages['NREL'],
                                                        geoid=geoid).first()
                # Transportation
                df.loc['Household Transportation Burden', (tract, 'EST')] = burden.transportation_value
                if burden.transportation_burden == 'Low':
                    df.loc['Household Transportation Burden', '% Points'] = 0
                elif burden.transportation_burden == 'Medium':
                    df.loc['Household Transportation Burden', '% Points'] = 0.5
                else:
                    df.loc['Household Transportation Burden', '% Points'] = 1

                # Energy
                df.loc['Household Energy Burden', (tract, 'EST')] = burden.energy_burden

                if tract == primary_tract:
                    if burden.energy_burden == 'Low':
                        df.loc['Household Energy Burden', '% Points'] = 0
                    elif burden.energy_burden == 'Medium':
                        df.loc['Household Energy Burden', '% Points'] = 0.5
                    else:
                        df.loc['Household Energy Burden', '% Points'] = 1

                df.loc['Household Transportation Burden', 'Source'] = 'NREL'
                df.loc['Household Energy Burden', 'Source'] = 'NREL'

            # Homelessness
            hud_states = list(Homelessness.objects.filter(dataset__vintage=vintages['HUD']).all())
            hud_data = []
            for hud_state in hud_states:
                homeless = hud_state.homeless
                population = hud_state.population
                per_capita = homeless / population
                hud_data.append([hud_state.state_fips, per_capita])

            quintile_df = pd.DataFrame(columns=['geoid', 'rate'], data=hud_data)
            quintile_df['quintile'] = pd.qcut(quintile_df['rate'], 5, labels=False)

            row = quintile_df.loc[quintile_df['geoid'] == state_fips]
            state_rank = quintile_df['rate'].rank().iloc[quintile_df.loc[quintile_df['geoid'] == state_fips].index[0]]
            df.loc['Homeless rate', (state_short_code, 'EST')] = int(state_rank)
            df.loc['Homeless rate', '% Points'] = (row.iloc[0].quintile * 0.25)
            df.loc['Homeless rate', 'Source'] = 'HUD'

        else:
            self.add_schools(vintages=vintages, state_fips=state.fips_code, district=school_district,
                             base_df=df)

        # State level life expectancy
        le_state = LifeExpectancy.objects.filter(geoid__exact=state.fips_code).first()
        df.loc['Life Expectancy', (state.short_code, 'EST')] = le_state.value
        df.loc['Life Expectancy', (state.short_code, 'SE')] = le_state.standard_error

        # State level CDC Places data (BRFSS, NCHS)
        brfss = list(BRFSS.objects.filter(dataset__vintage=vintages['BRFSS'], state__short_code=state.short_code))
        for b in brfss:
            df.loc[b.metric, (state.short_code, 'EST')] = b.value
            df.loc[b.metric, (state.short_code, 'MOE')] = b.moe

        # State level EJ Screen
        ejscreen = list(Metric.objects.filter(dataset__vintage=vintages['EJScreen'],
                                              dataset__data_source__name='EJ Screen', geoid=state.fips_code))
        for e in ejscreen:
            df.loc[e.name, (state.short_code, 'EST')] = e.value
            df.loc[e.name, (state.short_code, 'MOE')] = e.moe

        # Derivative values...Pandas does better with NaN, so we will convert our None values.
        df.replace(to_replace=[None], value=np.nan, inplace=True)

        # Aggregate MOE for All Tracts
        self.agg_moe(tracts=tracts, base_df=df)

        # Aggregate EST
        for index, row in df['All Tracts'].iterrows():
            if pd.isna(row['EST']):
                sums = 0
                for t in tracts:
                    if isinstance(df.loc[index, (t, 'EST')], numbers.Number):
                        sums += df.loc[index, (t, 'EST')]
                df.loc[index, ('All Tracts', 'EST')] = sums

        # Set up the columns that we will aggregate across...
        col = tracts.copy()
        col.extend(['All Tracts', state.short_code])

        # Aggregate Life Expectancy
        self.rollup_num_calc('Life Expectancy', 'Total Population', df, tracts, state, col)
        sums = 0
        for t in tracts:
            se = df.loc['Life Expectancy', (t, 'SE')]
            if np.isnan(se):
                continue
            else:
                sums += se ** 2
        df.loc['Life Expectancy', ('All Tracts', 'SE')] = sums**(1/2)
        df.loc['Life Expectancy', ('All Tracts', 'MOE')] = df.loc['Life Expectancy', ('All Tracts', 'SE')] * 1.645

        # Aggregate CDC Places
        df.loc['PLACES population', (['All Tracts', state.short_code], 'MOE')] = 0
        df.loc['PLACES population', (state.short_code, 'EST')] = None
        cdc_loader = CDCDataLoader(vintages['CDC'], self.user)
        for m in cdc_loader.metrics['CDC Places'].keys():
            self.rollup_percent_calc(m, 'PLACES population', df, tracts, state, col)

        # Aggregate EJ Screen
        ej_loader = EJScreenDataLoader(vintages['EJScreen'], self.user)
        for m in ej_loader.metrics['EJ Screen'].keys():
            self.row_avg(m, df, tracts)
            self.perc_avg(m, df, tracts)

        # Transit Use
        self.rollup_percent_calc('% Public Transit', 'Workers >16', df, tracts, state, col)
        self.rollup_percent_calc('% Walked', 'Workers >16', df, tracts, state, col)
        self.rollup_percent_calc('% Bicycle', 'Workers >16', df, tracts, state, col)

        # Female householders
        self.rollup_percent_calc('Female Householder (%)', 'Household Count', df, tracts, state, col)

        # Vehicle Miles
        self.rollup_num_calc('Average weekday vehicle miles traveled per household', 'Household Count', df, tracts,
                             state, col)

        # Median Household Income
        med_HH_income_scale = {'<10000': 0,
                               '10000-14999': 1,
                               '15000-24999': 2,
                               '25000-34999': 3,
                               '35000-49999': 4,
                               '50000-74999': 5,
                               '75000-99999': 6,
                               '100000-149999': 7,
                               '150000-199999': 8,
                               '>200000': 9}
        all_incomes = []

        # get total number of people in each income bracket
        for m in med_HH_income_scale.keys():
            temp = 'Total ' + m
            sums = 0
            for t in tracts:
                sums += df.loc[m, (t, 'EST')] * df.loc['Total with Income Data', (t, 'EST')] / 100
            df.loc[temp, ('All Tracts', 'EST')] = round(sums)

            # #add corresponding values to all_incomes
            num = df.loc[temp, ('All Tracts', 'EST')]
            arr = int(num) * [med_HH_income_scale[m]]
            all_incomes.extend(arr)

        med_scale = statistics.median(all_incomes)
        local_band = med_scale
        for m in med_HH_income_scale.keys():
            if med_scale == med_HH_income_scale[m]:
                med_range = m

        df.loc['Median Household Income', ('All Tracts', 'EST')] = med_scale
        df.astype({'All Tracts': object})

        # adjust the state number to be a band
        state_med = df.loc['Median Household Income', (state.short_code, 'EST')]
        for m in med_HH_income_scale.keys():
            if m == '<10000':
                if state_med <= 10000:
                    state_band = med_HH_income_scale[m]
                    state_range = m
                    df.loc['Median Household Income', (state.short_code, 'EST')] = state_band
            elif m == '>200000':
                if state_med >= 200000:
                    state_band = med_HH_income_scale[m]
                    state_range = m
                    df.at['Median Household Income', (state.short_code, 'EST')] = state_band
            else:
                split = m.split('-')
                low = float(split[0])
                high = float(split[1])
                if low <= state_med <= high:
                    state_band = med_HH_income_scale[m]
                    state_range = m
                    df.loc['Median Household Income', (state.short_code, 'EST')] = state_band


        self.divide_rows('Poverty Rate (%)', 'Below Poverty Level', 'Total with Poverty Data', 'Poverty Rate (%)',
                         df, col)
        self.divide_rows('Unemployment Rate (%)', 'Total Unemployed', 'In Labor Force', 'Unemployment Rate (%)',
                         df, col)
        self.divide_rows('Limited English-speaking Households (%)', 'Limited English-speaking',
                         'Total with Language Data', 'Limited English-speaking Households (%)', df, col)
        self.divide_rows('Population of Low-Income Children <5 (%)', 'Low-Income <5', 'Total <5',
                         'Population of Low-Income Children <5 (%)', df, col)
        self.divide_rows('Population of Low-Income Seniors >65 (%)', 'Low-Income >65', 'Total >65',
                         'Population of Low-Income Seniors >65 (%)', df, col)

        # Educational Attainment
        self.add_est('Total Educated', ['> 25 with Associates', '> 25 with Bachelors or higher'], df, col)
        self.divide_est('Associates, Bachelors degree, or higher (%)', 'Total Educated',
                        'Total with Education Data >25', df, col)
        self.add_moe('Total Educated', ['> 25 with Associates', '> 25 with Bachelors or higher'], df, col)
        self.divide_moe('Associates, Bachelors degree, or higher (%)', 'Total Educated',
                        'Total with Education Data >25', 'Associates, Bachelors degree, or higher (%)', df, col)

        # Population of Color
        df.loc['Total with Race Data', (state.short_code, 'MOE')] = 0
        self.subtract_est('Total Color', 'Total with Race Data', 'Total White Alone', df, col)
        self.divide_est('Population of Color (%)', 'Total Color', 'Total with Race Data', df, col)
        self.add_moe('Total Color', ['Total with Race Data', 'Total White Alone'], df, col)
        self.divide_moe('Population of Color (%)', 'Total Color', 'Total with Race Data', 'Population of Color (%)',
                        df, col)

        # Cost-Burdened Renters
        cbr_list = ['Rent 30.0-34.9%', 'Rent 35.0-39.9%', 'Rent 40.0-49.9%', 'Rent >50.0%']
        self.add_est('Total CBR', cbr_list, df, col)
        self.divide_est('Cost-Burdened Renters (%)', 'Total CBR', 'Total with Rent Data', df, col)
        self.add_moe('Total CBR', cbr_list, df, col)
        self.divide_moe('Cost-Burdened Renters (%)', 'Total CBR', 'Total with Rent Data',
                        'Cost-Burdened Renters (%)', df, col)

        # Car Ownership
        assign = (col, 'EST')
        df.loc['Average Number of Cars Per Household', assign] = \
            (df.loc['No vehicles', assign] * 0 + df.loc['1 vehicle', assign] * 1 +
             df.loc['2 vehicles', assign] * 2 + df.loc['>3 vehicles', assign] * 3) / \
            (df.loc['Occupied Housing Units', assign])
        df.loc['Average Number of Cars Per Household', 'Source'] = 'ACS'

        # Z SCORES
        i = 0
        valences = list(MetricValence.objects.filter(community__name='Advantaged'))
        for v in valences:
            df.loc[v.metric, 'Advantaged'] = True if v.valence > 0 else False

        i = 0
        valences = list(MetricValence.objects.filter(community__name='Disadvantaged'))
        for v in valences:
            df.loc[v.metric, 'Disadvantaged'] = True if v.valence > 0 else False

        for index, row in df['All Tracts'].iterrows():
            if df.loc[index, 'Source'].to_numpy() == 'PLACES':
                z = 1.96
            elif df.loc[index, 'Source'].to_numpy() == 'EPA':
                transit_freq = df.loc[index, (primary_tract, 'EST')]
                if np.isnan(transit_freq):
                    self.log.info(f"Defaulting to 50% for Transit Frequency because no data exists for primary tract={primary_tract}",
                                  extra={'user': self.user.username})
                    df.loc[index, '% Points'] = 0.5
                else:
                    # We need to apply the percentile tables from Emily's documentation, if we are in MA, CT, or RI.
                    # Otherwise, fall back to awarding 50% of points...Note that to use the percentiles we
                    # need to know the urban group, and here (as well as in the database...) that is represented as
                    # 1 == Urban, 2 == Suburban, 3 == Rural.
                    if self.is_NEF(scenario):
                        state_avg = SmartLocation.objects.filter(dataset__vintage=vintages['SmartLocation'],
                                                                 block_group_id__startswith=state_fips,
                                                                 D4c__gte=0).aggregate(Avg('D4c'))['D4c__avg']

                        df.loc[index, (state_short_code, 'EST')] = state_avg
                        ratio = transit_freq / state_avg if state_avg > 0 else 0
                        if ratio < 0.75:
                            df.loc[index, '% Points'] = 0
                        elif ratio < 1:
                            df.loc[index, '% Points'] = 0.25
                        elif ratio < 1.25:
                            df.loc[index, '% Points'] = 0.5
                        elif ratio < 2:
                            df.loc[index, '% Points'] = 0.75
                        else:
                            df.loc[index, '% Points'] = 1
                    else:
                        urbangroup = int(df.loc['Urban Group', (primary_tract, 'EST')])
                        percentile = ''
                        if state_short_code == 'MA':
                            if urbangroup == 1:
                                if transit_freq < 6:
                                    percentile = '0-9.99'
                                elif transit_freq < 44:
                                    percentile = '30-39.99'
                                elif transit_freq < 125:
                                    percentile = '40-49.99'
                                elif transit_freq < 225:
                                    percentile = '50-59.99'
                                elif transit_freq < 343:
                                    percentile = '60-69.99'
                                elif transit_freq < 497.4:
                                    percentile = '70-79.99'
                                elif transit_freq < 731.2:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'

                                df.loc[index, (primary_tract, 'PERC')] = percentile
                            elif urbangroup == 2:
                                if transit_freq < 1.33:
                                    percentile = '0-9.99'
                                elif transit_freq < 2.67:
                                    percentile = '60-69.99'
                                elif transit_freq < 10:
                                    percentile = '70-79.99'
                                elif transit_freq < 74.33:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'
                            else:
                                if transit_freq < 1.67:
                                    percentile = '0-9.99'
                                elif transit_freq < 3.33:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'
                        elif state_short_code == 'CT':
                            if urbangroup == 1:
                                if transit_freq < 0.33:
                                    percentile = '0-9.99'
                                else:
                                    percentile = '90-99.99'
                            elif urbangroup == 2:
                                if transit_freq == 0:
                                    percentile = '0-9.99'
                                else:
                                    percentile = '90-99.99'
                            else:
                                if transit_freq == 0:
                                    percentile = '0-9.99'
                                else:
                                    percentile = '90-99.99'

                            df.loc[index, (primary_tract, 'PERC')] = percentile
                        elif state_short_code == 'RI':
                            if urbangroup == 1:
                                if transit_freq < 27.4:
                                    percentile = '0-9.99'
                                elif transit_freq < 41.07:
                                    percentile = '10-19.99'
                                elif transit_freq < 52.1:
                                    percentile = '20-29.99'
                                elif transit_freq < 64.7:
                                    percentile = '30-39.99'
                                elif transit_freq < 75.3:
                                    percentile = '40-49.99'
                                elif transit_freq < 89.6:
                                    percentile = '50-59.99'
                                elif transit_freq < 108.47:
                                    percentile = '60-69.99'
                                elif transit_freq < 129.6:
                                    percentile = '70-79.99'
                                elif transit_freq < 174.23:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'
                            elif urbangroup == 2:
                                if transit_freq < 2.53:
                                    percentile = '0-9.99'
                                elif transit_freq < 9.27:
                                    percentile = '20-29.99'
                                elif transit_freq < 20.13:
                                    percentile = '30-39.99'
                                elif transit_freq < 31.17:
                                    percentile = '40-49.99'
                                elif transit_freq < 41.2:
                                    percentile = '50-59.99'
                                elif transit_freq < 52:
                                    percentile = '60-69.99'
                                elif transit_freq < 74.53:
                                    percentile = '70-79.99'
                                elif transit_freq < 96.47:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'
                            else:
                                if transit_freq < 2.67:
                                    percentile = '0-9.99'
                                elif transit_freq < 14:
                                    percentile = '50-59.99'
                                elif transit_freq < 32.33:
                                    percentile = '60-69.99'
                                elif transit_freq < 50.67:
                                    percentile = '70-79.99'
                                elif transit_freq < 91.5:
                                    percentile = '80-89.99'
                                else:
                                    percentile = '90-99.99'

                            df.loc[index, (primary_tract, 'PERC')] = percentile
                        else:
                            percentile = '50-59.99'
                            self.log.info(f"Defaulting to 50% for Transit Frequency because no percentile data exists for state={state_short_code}",
                                          extra={'user': self.user.username})

                        # Now assign % points, based on percentile...
                        if percentile.startswith('9'):
                            df.loc[index, '% Points'] = 1
                        elif percentile.startswith('8'):
                            df.loc[index, '% Points'] = 0.8
                        elif percentile.startswith('7'):
                            df.loc[index, '% Points'] = 0.7
                        elif percentile.startswith('6'):
                            df.loc[index, '% Points'] = 0.6
                        elif percentile.startswith('5'):
                            df.loc[index, '% Points'] = 0.5
                        else:
                            df.loc[index, '% Points'] = 0

            elif df.loc[index, 'Source'].to_numpy() == 'LATCH':
                if index == 'Average weekday vehicle miles traveled per household':
                    # To score points for the weekday miles metric, we need the state level data, which
                    # comes from our NHTS dataset.
                    nhts_data = NHTS.objects.filter(dataset__vintage=vintages['NHTS'],
                                                    fips_id=state_fips).first()

                    urban_groups = {1: 0, 2: 0, 3: 0}
                    for t in tracts:
                        urbangroup_value = df.loc['Urban Group', (t, 'EST')]
                        if not np.isnan(urbangroup_value):
                            urbangroup = int(urbangroup_value)
                            urban_groups[urbangroup] = urban_groups[urbangroup] + 1

                    weighted_avg = (urban_groups[1]*nhts_data.est_vmiles_urban +\
                                   urban_groups[2]*nhts_data.est_vmiles_suburban +\
                                   urban_groups[3]*nhts_data.est_vmiles_rural) / len(tracts)

                    ratio = df.loc[index, ('All Tracts', 'EST')] / weighted_avg
                    if ratio > 1:
                        df.loc[index, '% Points'] = 0
                    elif ratio > 0.9:
                        df.loc[index, '% Points'] = 0.5
                    elif ratio > 0.75:
                        df.loc[index, '% Points'] = 0.6
                    elif ratio > 0.5:
                        df.loc[index, '% Points'] = 0.8
                    else:
                        df.loc[index, '% Points'] = 1

            elif df.loc[index, 'Source'].to_numpy() == 'EDU':
                # Edu metrics use percentile to determine % points scored...
                if community == 'Advantaged':
                    perc = df.loc[index, ('All Tracts', 'PERC')]
                    if perc < 50:
                        df.loc[index, '% Points'] = 0
                    elif perc < 60:
                        df.loc[index, '% Points'] = 1/3
                    elif perc < 70:
                        df.loc[index, '% Points'] = 1/2
                    elif perc < 80:
                        df.loc[index, '% Points'] = 2/3
                    elif perc < 90:
                        df.loc[index, '% Points'] = 5/6
                    else:
                        df.loc[index, '% Points'] = 1

                continue

            elif df.loc[index, 'Source'].to_numpy() == 'EJ':
                # EJ metrics use percentile to determine % points scored...
                perc = df.loc[index, ('All Tracts', 'PERC')]
                if community == 'Disadvantaged':
                    if perc < 50:
                        df.loc[index, '% Points'] = 0
                    elif perc < 60:
                        df.loc[index, '% Points'] = 0.5
                    elif perc < 70:
                        df.loc[index, '% Points'] = 2/3
                    elif perc < 90:
                        df.loc[index, '% Points'] = 5/6
                    else:
                        df.loc[index, '% Points'] = 1
                elif community == 'Advantaged':
                    if perc > 50:
                        df.loc[index, '% Points'] = 0
                    elif perc > 40:
                        df.loc[index, '% Points'] = 0.5
                    elif perc > 30:
                        df.loc[index, '% Points'] = 2/3
                    elif perc > 10:
                        df.loc[index, '% Points'] = 5/6
                    else:
                        df.loc[index, '% Points'] = 1

                continue

            elif df.loc[index, 'Source'].to_numpy() == 'County Health':
                # Use Z Score logic but with county instead of 'All tracts'...
                z = 1.645
                z_score = 0
                index = 'Child Mortality'
                X1 = df.loc[index, (county_name, 'EST')] / 100
                X2 = df.loc[index, (state.short_code, 'EST')] / 100
                SE1 = (df.loc[index, (county_name, 'MOE')] / 100) / z
                SE2 = (df.loc[index, (state.short_code, 'MOE')] / 100) / z
                if (SE1**2 + SE2**2)**(1/2) == 0:
                    pass
                else:
                    df.loc[index, 'Z Score'] = np.real((X1 - X2) / (SE1**2 + SE2**2)**(1/2))

                z_score = df.loc[index, 'Z Score'].values

                if abs(z_score) < 1.282:
                    df.loc[index, '% Points'] = 0.5
                elif community == 'Disadvantaged':
                    if z_score < -1.282:
                        df.loc[index, '% Points'] = 0
                    elif 1.282 < z_score < 1.645:
                        df.loc[index, '% Points'] = 2/3
                    elif 1.645 < z_score < 2.326:
                        df.loc[index, '% Points'] = 5/6
                    elif z_score > 2.326:
                        df.loc[index, '% Points'] = 1
                else:
                    if z_score > 1.282:
                        df.loc[index, '% Points'] = 0
                    elif -1.282 > z_score > -1.645:
                        df.loc[index, '% Points'] = 2/3
                    elif -1.645 > z_score > -2.326:
                        df.loc[index, '% Points'] = 5/6
                    elif z_score < -2.326:
                        df.loc[index, '% Points'] = 1
            else:
                z = 1.645

            z_score = 0

            if index == 'Life Expectancy':
                X1 = df.loc[index, ('All Tracts', 'EST')] / 100
                X2 = df.loc[index, (state.short_code, 'EST')] / 100
                SE1 = df.loc[index, ('All Tracts', 'SE')]
                SE2 = df.loc[index, (state.short_code, 'SE')]
                if (SE1**2 + SE2**2)**(1/2) == 0:
                    pass
                else:
                    df.loc[index, 'Z Score'] = np.real((X1 - X2) / (SE1**2 + SE2**2)**(1/2))
                z_score = df.loc[index, 'Z Score'].values
            elif index != 'Median Household Income' and index != 'Child Mortality':
                X1 = df.loc[index, ('All Tracts', 'EST')] / 100
                X2 = df.loc[index, (state.short_code, 'EST')] / 100
                SE1 = (df.loc[index, ('All Tracts', 'MOE')] / 100) / z
                SE2 = (df.loc[index, (state.short_code, 'MOE')] / 100) / z
                if (SE1**2 + SE2**2)**(1/2) == 0:
                    pass
                else:
                    df.loc[index, 'Z Score'] = np.real((X1 - X2) / (SE1**2 + SE2**2)**(1/2))
                z_score = df.loc[index, 'Z Score'].values

            # Special case...median income does not use Z Score
            if index == 'Median Household Income':
                if community == 'Disadvantaged':
                    difference = state_band - local_band
                    if difference <= 0:
                        df.loc[index, '% Points'] = 0
                    elif difference == 1:
                        df.loc[index, '% Points'] = 1/2
                    elif difference == 2:
                        df.loc[index, '% Points'] = 2/3
                    else:
                        df.loc[index, '% Points'] = 1
                else:
                    difference = local_band - state_band
                    if difference <= 0:
                        df.loc[index, '% Points'] = 0
                    elif difference == 1:
                        df.loc[index, '% Points'] = 1/2
                    elif difference == 2:
                        df.loc[index, '% Points'] = 2/3
                    else:
                        df.loc[index, '% Points'] = 1

            elif index == 'Child Mortality':
                pass
            elif abs(z_score) < 1.282:
                df.loc[index, '% Points'] = 0.5
            elif df.loc[index, community].values:
                if z_score < -1.282:
                    df.loc[index, '% Points'] = 0
                elif 1.282 < z_score < 1.645:
                    df.loc[index, '% Points'] = 2/3
                elif 1.645 < z_score < 2.326:
                    df.loc[index, '% Points'] = 5/6
                elif z_score > 2.326:
                    df.loc[index, '% Points'] = 1

            elif not df.loc[index, community].values:
                if z_score > 1.282:
                    df.loc[index, '% Points'] = 0
                elif -1.282 > z_score > -1.645:
                    df.loc[index, '% Points'] = 2/3
                elif -1.645 > z_score > -2.326:
                    df.loc[index, '% Points'] = 5/6
                elif z_score < -2.326:
                    df.loc[index, '% Points'] = 1

        df.loc['Average weekday vehicle miles (state urban)', (state.short_code, 'EST')] \
            = nhts_data.est_vmiles_urban
        df.loc['Average weekday vehicle miles (state suburban)', (state.short_code, 'EST')] \
            = nhts_data.est_vmiles_suburban
        df.loc['Average weekday vehicle miles (state rural)', (state.short_code, 'EST')] \
            = nhts_data.est_vmiles_rural

        # Urban Group
        primary_group = df.loc['Urban Group', (primary_tract, 'EST')]
        df.loc['Urban Group', ('All Tracts', 'EST')] = primary_group
        if primary_group == 1:
            df.loc['Urban Group', ('All Tracts', 'EST')] = 'URB'
        elif primary_group == 2:
            df.loc['Urban Group', ('All Tracts', 'EST')] = 'SUB'
        elif primary_group == 3:
            df.loc['Urban Group', ('All Tracts', 'EST')] = 'RUR'

        # Ratio
        df['Ratio'] = df['All Tracts']['EST'] / df[state.short_code]['EST']

        # Rename points scored column and make sure it's in the last position
        new_name = '% of Max Points Scored (Decimal Value)'
        df = df.rename({'% Points': new_name}, axis='columns')
        points_column = df.pop(new_name)
        df.insert(len(df.columns), new_name, points_column)

        df.at['Median Household Income', 'Ratio'] = df.at['Median Household Income', ('All Tracts', 'EST')] - df.at['Median Household Income', (state.short_code, 'EST')]

        # Now that all calculations are done, add the income range to median household income
        df.loc['Median Household Income', ('All Tracts', 'EST')] = f'{med_range} (band {local_band})'
        df.loc['Median Household Income', (state.short_code, 'EST')] = f'{state_range} (band {state_band})'

        # Create the scorecard sheet, which is a summary of selected metrics from the raw data
        if community == 'Disadvantaged':
            row_order = ['Life Expectancy',
                         'Cancer (excluding skin cancer) among adults >= 18',
                         'Current asthma among adults >= 18',
                         'COPD among adults >= 18',
                         'Coronary heart disease among adults >= 18',
                         'Diabetes among adults >= 18',
                         'Stroke among adults >= 18',
                         'Mental health not good for >= 14 days among adults >= 18',
                         'PM 2.5 (ug/m3)',
                         'NATA Diesel PM (ug/m3)',
                         'NATA Air Toxics Cancer Risk (risk per MM)',
                         'NATA Respiratory Hazard Index',
                         '% Public Transit',
                         '% Walked',
                         '% Bicycle',
                         'Average weekday vehicle miles traveled per household',
                         'Average weekday vehicle miles (state urban)',
                         'Average weekday vehicle miles (state suburban)',
                         'Average weekday vehicle miles (state rural)',
                         'Urban Group',
                         'Transit Frequency',
                         'Median Household Income',
                         'Poverty Rate (%)',
                         'Unemployment Rate (%)',
                         'Associates, Bachelors degree, or higher (%)',
                         'Limited English-speaking Households (%)',
                         'Population of Color (%)',
                         'Cost-Burdened Renters (%)',
                         'Population of Low-Income Children <5 (%)',
                         'Population of Low-Income Seniors >65 (%)',
                         'Average Number of Cars Per Household']
        elif community == 'Advantaged':
            row_order = ['Life Expectancy',
                         'Cancer (excluding skin cancer) among adults >= 18',
                         'Current asthma among adults >= 18',
                         'COPD among adults >= 18',
                         'Coronary heart disease among adults >= 18',
                         'Diabetes among adults >= 18',
                         'Stroke among adults >= 18',
                         'Mental health not good for >= 14 days among adults >= 18',
                         'PM 2.5 (ug/m3)',
                         'NATA Diesel PM (ug/m3)',
                         'NATA Air Toxics Cancer Risk (risk per MM)',
                         'NATA Respiratory Hazard Index',
                         '% Public Transit',
                         '% Walked',
                         '% Bicycle',
                         'Average weekday vehicle miles traveled per household',
                         'Average weekday vehicle miles (state urban)',
                         'Average weekday vehicle miles (state suburban)',
                         'Average weekday vehicle miles (state rural)',
                         'Urban Group',
                         'Transit Frequency',
                         'Median Household Income',
                         'Poverty Rate (%)',
                         'Unemployment Rate (%)',
                         'Associates, Bachelors degree, or higher (%)',
                         'Cost-Burdened Renters (%)',
                         'Population of Low-Income Children <5 (%)',
                         'Population of Low-Income Seniors >65 (%)',
                         'School Performance - Overall',
                         'School Performance - Econ. Disadvantaged',
                         'Average Number of Cars Per Household']

        # Some extra metrics for the national scenario...
        if self.is_NEF(scenario):
            row_order.insert(row_order.index('Average Number of Cars Per Household'), 'Female Householder (%)')
            row_order.insert(row_order.index('Average Number of Cars Per Household')+1, 'Child Mortality')
            row_order.insert(row_order.index('Child Mortality')+1, 'Total healthcare cost per capita per household income')
            row_order.insert(row_order.index('Total healthcare cost per capita per household income')+1, 'Child obesity')
            row_order.insert(row_order.index('Child obesity')+1, 'Child asthma')
            row_order.insert(row_order.index('Child asthma')+1, 'ER Visits')
            row_order.insert(row_order.index('ER Visits')+1, 'Maternity Care')
            row_order.insert(row_order.index('Maternity Care')+1, 'Access to Mental Health Care')
            row_order.insert(row_order.index('Access to Mental Health Care')+1, 'Household Transportation Burden')
            row_order.insert(row_order.index('Household Transportation Burden')+1, 'Household Energy Burden')
            row_order.insert(row_order.index('Household Energy Burden')+1, 'Homeless rate')
            row_order.insert(row_order.index('Homeless rate')+1, 'Population With Disability (%)')


        scorecard_df = df.reindex(row_order)
        scorecard_df.to_excel(writer, sheet_name='Scorecard')

        # Create the raw data sheet
        if include_raw:
            df.to_excel(writer, sheet_name='Raw data')

        # Create the appendix sheet that describes the dataset vintages and parameters used
        time = datetime.now(tz=EST5EDT()).strftime("%I:%M%p on %B %d, %Y")
        vintages = DataService.vintages()

        ds = Dataset.objects.filter(data_source__name__exact='ACS Detail', vintage=vintages['ACS']).first()
        acs_detail_ref = ds.data_source.reference_url
        acs_detail_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='ACS Subject', vintage=vintages['ACS']).first()
        acs_subject_ref = ds.data_source.reference_url
        acs_subject_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='ACS Profile', vintage=vintages['ACS']).first()
        acs_profile_ref = ds.data_source.reference_url
        acs_profile_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='CDC Places', vintage=vintages['CDC']).first()
        cdc_ref = ds.data_source.reference_url
        cdc_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='EJ Screen', vintage=vintages['EJScreen']).first()
        ej_ref = ds.data_source.reference_url
        ej_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='USALEEP', vintage=vintages['LifeExpectancy']).first()
        life_ref = ds.data_source.reference_url
        life_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='LATCH', vintage=vintages['Latch']).first()
        latch_ref = ds.data_source.reference_url
        latch_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='SmartLocation', vintage=vintages['SmartLocation']).first()
        smart_ref = ds.data_source.reference_url
        smart_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='BRFSS', vintage=vintages['BRFSS']).first()
        brfss_ref = ds.data_source.reference_url
        brfss_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='NMTC', vintage=vintages['NMTC']).first()
        nmtc_ref = ds.data_source.reference_url
        nmtc_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='OpportunityZone', vintage=vintages['OpportunityZone']).first()
        zones_ref = ds.data_source.reference_url
        zones_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='NHTS', vintage=vintages['NHTS']).first()
        nhts_ref = ds.data_source.reference_url
        nhts_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='Education MA', vintage=vintages['EducationMA']).first()
        eduma_ref = ds.data_source.reference_url
        eduma_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='Education MA Subgroup', vintage=vintages['EducationMA Subgroup']).first()
        edumasub_ref = ds.data_source.reference_url
        edumasub_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='Education CT', vintage=vintages['EducationCT']).first()
        educt_ref = ds.data_source.reference_url
        educt_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='Education CT Subgroup', vintage=vintages['EducationCT Subgroup']).first()
        eductsub_ref = ds.data_source.reference_url
        eductsub_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='Education RI', vintage=vintages['EducationRI']).first()
        eduri_ref = ds.data_source.reference_url
        eduri_desc = ds.descriptor

        ds = Dataset.objects.filter(data_source__name__exact='School District', vintage=vintages['SchoolDistrict']).first()
        dist_ref = ds.data_source.reference_url
        dist_desc = ds.descriptor

        if self.is_NEF(scenario):

            ds = Dataset.objects.filter(data_source__name__exact='County Health Rankings', vintage=vintages['ChildMortality']).first()
            cm_ref = ds.data_source.reference_url
            cm_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='CMS', vintage=vintages['PersonalHealthCare']).first()
            phc_ref = ds.data_source.reference_url
            phc_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='NSCH', vintage=vintages['NSCH']).first()
            nsch_ref = ds.data_source.reference_url
            nsch_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='KFF', vintage=vintages['ERVisits']).first()
            kff_ref = ds.data_source.reference_url
            kff_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='MoD', vintage=vintages['MaternityCare']).first()
            mod_ref = ds.data_source.reference_url
            mod_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='MHA', vintage=vintages['MentalHealthCare']).first()
            mha_ref = ds.data_source.reference_url
            mha_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='NREL', vintage=vintages['NREL']).first()
            nrel_ref = ds.data_source.reference_url
            nrel_desc = ds.descriptor

            ds = Dataset.objects.filter(data_source__name__exact='HUD', vintage=vintages['HUD']).first()
            hud_ref = ds.data_source.reference_url
            hud_desc = ds.descriptor

        data = {f"Created by {user} at {time}": ['', '', ''],
                'Address': [address, '', ''],
                'State FIPS code': [state_fips, '', ''],
                'State': [state.short_code, '', ''],
                'County code': [county, '', ''],
                'County name': [county_name, '', ''],
                'All selected census tracts': [tracts, '', ''],
                'Primary census tract': [primary_tract, '', ''],
                'Census block group': [block, '', ''],
                'School district': [school_district, '', ''],
                'Community type': [community, '', ''],
                'Buffer radius': [buffer_radius, '', ''],
                'Population threshold': [population_threshold, '', ''],
                '': ['', '', ''],
                'Datasets': ['', '', ''],
                'ACS Detail': [f"Vintage: {vintages['ACS']}", f"{acs_detail_desc}", f"{acs_detail_ref}"],
                'ACS Subject': [f"Vintage: {vintages['ACS']}", f"{acs_subject_desc}", f"{acs_subject_ref}"],
                'ACS Profile': [f"Vintage: {vintages['ACS']}", f"{acs_profile_desc}", f"{acs_profile_ref}"],
                'CDC': [f"Vintage: {vintages['CDC']} ", f"{cdc_desc}", f"{cdc_ref}"],
                'EJScreen': [f"Vintage: {vintages['EJScreen']}", f"{ej_desc}", f"{ej_ref}"],
                'Life Expectancy': [f"Vintage: {vintages['LifeExpectancy']}", f"{life_desc}", f"{life_ref}"],
                'Latch': [f"Vintage: {vintages['Latch']}", f"{latch_desc}", f"{latch_ref}"],
                'Smart Location': [f"Vintage: {vintages['SmartLocation']}", f"{smart_desc}", f"{smart_ref}"],
                'BRFSS': [f"Vintage: {vintages['BRFSS']}", f"{brfss_desc}", f"{brfss_ref}"],
                'NMTC': [f"Vintage: {vintages['NMTC']}", f"{nmtc_desc}", f"{nmtc_ref}"],
                'Opportunity Zones': [f"Vintage: {vintages['OpportunityZone']}", f"{zones_desc}", f"{zones_ref}"],
                'NHTS': [f"Vintage: {vintages['NHTS']}", f"{nhts_desc}", f"{nhts_ref}"],
                'Education MA': [f"Vintage: {vintages['EducationMA']}", f"{eduma_desc}", f"{eduma_ref}"],
                'Education MA Subgroup': [f"Vintage: {vintages['EducationMA Subgroup']}", f"{edumasub_desc}", f"{edumasub_ref}"],
                'Education CT': [f"Vintage: {vintages['EducationCT']}", f"{educt_desc}", f"{educt_ref}"],
                'Education CT Subgroup': [f"Vintage: {vintages['EducationCT Subgroup']}", f"{eductsub_desc}", f"{eductsub_ref}"],
                'Education RI': [f"Vintage: {vintages['EducationRI']}", f"{eduri_desc}", f"{eduri_ref}"],
                'School District': [f"Vintage: {vintages['SchoolDistrict']}", f"{dist_desc}", f"{dist_ref}"]
                }

        if self.is_NEF(scenario):
            data['ChildMortality'] = [f"Vintage: {vintages['ChildMortality']}", f"{cm_desc}", f"{cm_ref}"]
            data['PersonalHealthCare'] = [f"Vintage: {vintages['PersonalHealthCare']}", f"{phc_desc}", f"{phc_ref}"]
            data['NSCH'] = [f"Vintage: {vintages['NSCH']}", f"{nsch_desc}", f"{nsch_ref}"]
            data['KFF'] = [f"Vintage: {vintages['ERVisits']}", f"{kff_desc}", f"{kff_ref}"]
            data['MoD'] = [f"Vintage: {vintages['MaternityCare']}", f"{mod_desc}", f"{mod_ref}"]
            data['MHA'] = [f"Vintage: {vintages['MentalHealthCare']}", f"{mha_desc}", f"{mha_ref}"]
            data['NREL'] = [f"Vintage: {vintages['NREL']}", f"{nrel_desc}", f"{nrel_ref}"]
            data['HUD'] = [f"Vintage: {vintages['HUD']}", f"{hud_desc}", f"{hud_ref}"]

        appendix_df = pd.DataFrame.from_dict(data, orient='index', columns=['Parameter', 'Descriptor', 'Reference URL'])
        appendix_df.to_excel(writer, sheet_name='Appendix')

        # Perform any required formatting on the spreadsheet(s)
        workbook = writer.book
        # decimal_format = workbook.add_format({'num_format': '0.000'})

        # Create the housing affordability sheet (if necessary)
        if housing_data is not None and len(housing_data) > 0 and housing_data[0]['unit_type'] != '':
            worksheet = workbook.create_sheet('Housing Affordability')

            worksheet.column_dimensions['A'].width = 24
            worksheet.column_dimensions['B'].width = 10
            worksheet.column_dimensions['D'].width = 11
            worksheet.cell(1, 1).value = 'Type'
            worksheet.cell(1, 2).value = 'Number'
            worksheet.cell(1, 3).value = '% of Total'
            worksheet.cell(1, 4).value = 'Starting Rent'

            row = 2
            total_count = sum(int(hd['unit_count']) for hd in housing_data)
            total_weighted_rent = sum(int(hd['unit_count']) * int(hd['starting_rent']) for hd in housing_data) / total_count
            for hd in housing_data:
                count = int(hd['unit_count'])
                rent = int(hd['starting_rent'])
                worksheet.cell(row, 1).value = hd['unit_type']
                worksheet.cell(row, 2).value = count
                worksheet.cell(row, 3).value = count / total_count
                worksheet.cell(row, 3).number_format = '0%'
                worksheet.cell(row, 4).value = rent
                worksheet.cell(row, 4).number_format = '$#,##0'
                row = row + 1

            row = row + 2
            worksheet.cell(row, 1).value = 'Total Units'
            worksheet.cell(row, 2).value = total_count

            row = row + 2
            worksheet.cell(row, 1).value = 'Weighted avg. rental cost'
            worksheet.cell(row, 2).value = total_weighted_rent
            worksheet.cell(row, 2).number_format = '$#,##0.00'

            income_needed = total_weighted_rent * 12 / 0.3
            row = row + 2
            worksheet.cell(row, 1).value = 'Annual income needed'
            worksheet.cell(row, 2).value = income_needed
            worksheet.cell(row, 2).number_format = '$#,##0.00'

            row = row + 2

            # Let's get the income data from our raw dataset. First slice the columns we want...
            tract_list = tracts + ['All Tracts']
            idx = pd.IndexSlice
            slice_df = df.loc[:, idx[tract_list, :]]

            # Then winnow the rows...
            start_index = slice_df.index.get_loc('Total with Income Data')
            end_index = start_index + 11
            slice_df = slice_df.iloc[start_index:end_index, :]

            grand_total_who_can_afford = 0
            for tract in tracts:
                if income_needed < 10000:
                    slice_df.loc['<10000>', (tract, 'SE')] = slice_df[tract]['EST']['<10000'] *\
                                                            slice_df[tract]['EST']['Total with Income Data']
                if income_needed < 15000:
                    slice_df.loc['10000-14999', (tract, 'SE')] = slice_df[tract]['EST']['10000-14999'] * \
                                                             slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 25000:
                    slice_df.loc['15000-24999', (tract, 'SE')] = slice_df[tract]['EST']['15000-24999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 35000:
                    slice_df.loc['25000-34999', (tract, 'SE')] = slice_df[tract]['EST']['25000-34999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 50000:
                    slice_df.loc['35000-49999', (tract, 'SE')] = slice_df[tract]['EST']['35000-49999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 75000:
                    slice_df.loc['50000-74999', (tract, 'SE')] = slice_df[tract]['EST']['50000-74999'] * \
                                                            slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 100000:
                    slice_df.loc['75000-99999', (tract, 'SE')] = slice_df[tract]['EST']['75000-99999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 150000:
                    slice_df.loc['100000-149999', (tract, 'SE')] = slice_df[tract]['EST']['100000-149999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                if income_needed < 200000:
                    slice_df.loc['150000-199999', (tract, 'SE')] = slice_df[tract]['EST']['150000-199999'] * \
                                                                 slice_df[tract]['EST']['Total with Income Data'] / 100
                slice_df.loc['>200000', (tract, 'SE')] = slice_df[tract]['EST']['>200000'] * \
                                                             slice_df[tract]['EST']['Total with Income Data'] / 100

                total_who_can_afford = slice_df.loc[:, (tract, 'SE')].sum()
                slice_df.loc['Total who can afford', (tract, 'SE')] = total_who_can_afford
                grand_total_who_can_afford = grand_total_who_can_afford + total_who_can_afford

            slice_df.loc['Total who can afford', ('All Tracts', 'SE')] = grand_total_who_can_afford
            percent_afford = round(100 * grand_total_who_can_afford / slice_df['All Tracts']['EST']['Total with Income Data'])

            points_scored = 0
            if percent_afford < 25:
                points_scored = 0
            elif percent_afford < 50:
                points_scored = 0.5
            elif percent_afford < 75:
                points_scored = 0.75
            else:
                points_scored = 1

            slice_df.loc['Percent who can afford', ('All Tracts', 'SE')] = str(percent_afford) + '%'
            slice_df.loc['% Points scored', ('All Tracts', 'SE')] = points_scored
            slice_df = slice_df.rename(columns={'EST': '% People', 'SE': '# People', 'PERC': ''}, level=1)

            for ws_row in worksheet.iter_rows(min_row=int(15+len(housing_data)), min_col=2, max_col=100, max_row=150):
                for cell in ws_row:
                    cell.number_format = '0.000'
            slice_df.to_excel(writer, sheet_name='Housing Affordability', startrow=row)

        # Scorecard...
        worksheet = writer.sheets['Scorecard']
        worksheet.column_dimensions['A'].width = 60
        for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=100, max_row=150):
            for cell in row:
                cell.number_format = '0.000'

        # Raw data...
        if include_raw:
            worksheet = writer.sheets['Raw data']
            worksheet.column_dimensions['A'].width = 60
            for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=100, max_row=150):
                for cell in row:
                    cell.number_format = '0.000'

        # Appendix...
        worksheet = writer.sheets['Appendix']
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 60
        worksheet.column_dimensions['D'].width = 60

        writer.save()
        xlsx_data = output.getvalue()

        return xlsx_data

    def construct_dataframe(self, state: str, tracts: list, county_name: str, scenario: str):
        col = []
        for t in tracts:
            col.append(t)

        # MAKE COLUMNS
        temp = ['All Tracts', county_name, state, 'US'] if self.is_NEF(scenario) else ['All Tracts', state]
        col.extend(temp)

        # Second level
        col2 = ["EST", "SE", "MOE", "PERC"]
        iterables = [col, col2]
        index = pd.MultiIndex.from_product(iterables=iterables)

        # MAKE ROWS
        ind = ['Life Expectancy',
               'Cancer (excluding skin cancer) among adults >= 18',
               'Current asthma among adults >= 18',
               'COPD among adults >= 18',
               'Coronary heart disease among adults >= 18',
               'Diabetes among adults >= 18',
               'Stroke among adults >= 18',
               'Mental health not good for >= 14 days among adults >= 18',
               'PM 2.5 (ug/m3)',
               'NATA Diesel PM (ug/m3)',
               'NATA Air Toxics Cancer Risk (risk per MM)',
               'NATA Respiratory Hazard Index',
               '% Public Transit',
               '% Walked',
               '% Bicycle',
               'Average weekday vehicle miles traveled per household',
               'Average weekday vehicle miles (state urban)',
               'Average weekday vehicle miles (state suburban)',
               'Average weekday vehicle miles (state rural)',
               'Urban Group',
               'Transit Frequency',
               'Median Household Income',
               'Poverty Rate (%)',
               'Unemployment Rate (%)',
               'Associates, Bachelors degree, or higher (%)',
               'Limited English-speaking Households (%)',
               'Population of Color (%)',
               'Cost-Burdened Renters (%)',
               'Population of Low-Income Children <5 (%)',
               'Population of Low-Income Seniors >65 (%)',
               'School Performance - Overall',
               'School Performance - Econ. Disadvantaged',
               'Average Number of Cars Per Household']

        # MAKE THE DATAFRAME
        an_array = np.empty((len(ind),len(col)*len(col2)))
        an_array[:] = np.NaN
        return pd.DataFrame((an_array), index=ind, columns=index, dtype='object')

    def add_acs_all(self, vintage: str, state: State, county: str, tract: str, base_df: DataFrame):

        tract_full = state.fips_code + county + tract
        data_loader = ACSDataLoader(vintage, self.user)
        for ds in data_loader.metrics.keys():
            for m in data_loader.metrics[ds].keys():
                try:
                    est = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full, name=m).get()
                    moe = Metric.objects.values_list('moe', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full, name=m).get()

                    base_df.loc[m, (tract, 'EST')] = est
                    base_df.loc[m, (tract, 'MOE')] = moe

                    # state data...
                    est = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=state.fips_code, name=m).get()
                    moe = Metric.objects.values_list('moe', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=state.fips_code, name=m).get()

                    base_df.loc[m, (state.short_code, 'EST')] = est
                    base_df.loc[m, (state.short_code, 'MOE')] = moe
                    base_df.loc[m, 'Source'] = 'ACS'
                except ObjectDoesNotExist:
                    self.log.error(f"Couldn't find metric {m}")
                except KeyError:
                    continue

    def add_cdc_tract(self, vintage: str, state_fips: str, county: str, tract: str, base_df: DataFrame):

        tract_full = state_fips + county + tract
        data_loader = CDCDataLoader(vintage, self.user)
        for ds in data_loader.metrics.keys():
            for m in data_loader.metrics[ds].keys():
                try:
                    est = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full, name=m).get()
                    moe = Metric.objects.values_list('moe', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full, name=m).get()
                    pop = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full,
                        name='PLACES Population').get()

                    base_df.loc[m, (tract, 'EST')] = est
                    base_df.loc[m, (tract, 'MOE')] = moe
                    base_df.loc['PLACES population', (tract, 'EST')] = pop
                    base_df.loc['PLACES population', (tract, 'MOE')] = 0

                    base_df.loc[m, 'Source'] = 'PLACES'
                except ObjectDoesNotExist as e:
                    self.log.info(f"No CDC metrics found for metric={m}, geoid={tract_full}, vintage={vintage}",
                                  extra={'user': self.user.username})
                    continue
                except KeyError:
                    continue

    def add_ejscreen_tract(self, vintage: str, state_fips: str, county: str, tract: str, base_df: DataFrame):

        tract_full = state_fips + county + tract
        data_loader = EJScreenDataLoader(vintage, self.user)
        for ds in data_loader.metrics.keys():
            for m in data_loader.metrics[ds].keys():
                try:
                    est = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full, name=m).get()
                    perc = Metric.objects.values_list('value', flat=True).filter(
                        dataset__data_source__name=ds, dataset__vintage=vintage, geoid=tract_full,
                        name='Percentile ' + m).get()

                    base_df.loc[m, (tract, 'EST')] = est
                    base_df.loc[m, (tract, 'PERC')] = perc
                    base_df.loc[m, 'Source'] = 'EJ'
                except KeyError:
                    continue
                except ObjectDoesNotExist:
                    print("Missing EJScreen data...API may be broken!")
                    continue

    def add_life_exp_tract(self, vintage: str, state_fips: str, county: str, tract: str, base_df: DataFrame):
        metric = 'Life Expectancy'
        tract_full = state_fips + county + tract

        try:
            est = LifeExpectancy.objects.values_list('value', flat=True).filter(
                dataset__vintage=vintage, geoid=tract_full).get()
            se = LifeExpectancy.objects.values_list('standard_error', flat=True).filter(
                dataset__vintage=vintage, geoid=tract_full).get()

            base_df.loc[metric, (tract, 'EST')] = est
            base_df.loc[metric, (tract, 'SE')] = se
            base_df.loc[metric, 'Source'] = 'RWJF'
        except ObjectDoesNotExist as e:
            self.log.info(f"No Life Expectancy metric found for geoid={tract_full}, vintage={vintage}",
                          extra={'user': self.user.username})

    def add_latch_tract(self, vintage: str, state_fips: str, county: str, tract: str, base_df: DataFrame):

        tract_full = state_fips + county + tract

        try:
            urban_group = Latch.objects.values_list('urban_group', flat=True).filter(
                dataset__vintage=vintage, tract_id=tract_full).get()
            est_vmiles = Latch.objects.values_list('est_vmiles', flat=True).filter(
                dataset__vintage=vintage, tract_id=tract_full).get()
            hh_cnt = Latch.objects.values_list('hh_cnt', flat=True).filter(
                dataset__vintage=vintage, tract_id=tract_full).get()

            base_df.loc['Urban Group', (tract, 'EST')] = urban_group
            base_df.loc['Urban Group', 'Source'] = 'LATCH'

            base_df.loc['Average weekday vehicle miles traveled per household', (tract, 'EST')] = est_vmiles
            base_df.loc['Average weekday vehicle miles traveled per household', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state urban)', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state suburban)', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state rural)', 'Source'] = 'LATCH'

            base_df.loc['Household Count', (tract, 'EST')] = hh_cnt
            base_df.loc['Household Count', 'Source'] = 'LATCH'
        except ObjectDoesNotExist as e:
            self.log.info(f"No LATCH metrics found for geoid={tract_full}, vintage={vintage}",
                          extra={'user': self.user.username})
            base_df.loc['Average weekday vehicle miles traveled per household', (tract, 'EST')] = None
            base_df.loc['Average weekday vehicle miles traveled per household', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state urban)', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state suburban)', 'Source'] = 'LATCH'
            base_df.loc['Average weekday vehicle miles (state rural)', 'Source'] = 'LATCH'

    def add_smart_location_all(self, vintage: str, state_fips: str, county: str, primary: str, block: str,
                               base_df: DataFrame):

        geoid = state_fips + county + primary + block

        base_df.loc['Transit Frequency', 'Source'] = 'EPA'

        try:
            transit_freq = SmartLocation.objects.values_list('D4c', flat=True).filter(
                dataset__vintage=vintage, block_group_id=geoid).get()

            base_df.loc['Transit Frequency', (primary, 'EST')] = transit_freq

        except ObjectDoesNotExist as e:
            base_df.loc['Transit Frequency', (primary, 'EST')] = np.nan
            self.log.info(f"No SmartLocation metrics found for geoid={geoid}, vintage={vintage}",
                          extra={'user': self.user.username})

    def add_schools(self, vintages: list, state_fips: str, district: str, base_df:DataFrame):
        # MA
        if state_fips == '25':

            try:
                # school performance
                district_avg = EducationMA.objects.filter(dataset__vintage=vintages['EducationMA'], district=district).aggregate(Avg('percentile'))
                base_df.loc['School Performance - Overall', ('All Tracts', 'PERC')] = district_avg['percentile__avg']

                # disadvantaged
                percentiles = []
                schools = []

                # Prioritize Econ. Disadvantaged over High needs group
                disadvantaged = list(
                    EducationMASubgroup.objects.filter(dataset__vintage=vintages['EducationMA Subgroup'], district=district, group='Econ. Disadvantaged'))
                for d in disadvantaged:
                    schools.append(d.school)
                    percentiles.append(d.percentile)

                high_needs = list(
                    EducationMASubgroup.objects.filter(dataset__vintage=vintages['EducationMA Subgroup'], district=district, group='High needs'))
                for hn in high_needs:
                    if hn.school not in schools:
                        percentiles.append(hn.percentile)

                base_df.at['School Performance - Econ. Disadvantaged', ('All Tracts', 'PERC')] = \
                    np.sum(percentiles) / len(percentiles)

                base_df.loc['School Performance - Overall', 'Source'] = 'EDU'
                base_df.loc['School Performance - Econ. Disadvantaged', 'Source'] = 'EDU'
            except BaseException:
                print("Error creating spreadsheet for MA schools.")

        # CT
        if state_fips == '09':

            try:
                # school performance
                district_avg = EducationCT.objects.filter(dataset__vintage=vintages['EducationCT'], district=district).aggregate(Avg('percentile'))
                base_df.at['School Performance - Overall', ('All Tracts', 'PERC')] = district_avg['percentile__avg']

                # high needs
                all_districts = list(EducationCTSubgroup.objects.filter(dataset__vintage=vintages['EducationCT Subgroup']).values('district').distinct())
                all_scores = []
                for d in all_districts:
                    subject_scores = []

                    edu_hn = EducationCTSubgroup.objects.filter(dataset__vintage=vintages['EducationCT Subgroup'], district=d['district'], subgroup='High Needs').first()

                    if edu_hn is None:
                        continue

                    ela = edu_hn.ela_performance_index
                    if ela is not None:
                        subject_scores.append(ela)
                    math = edu_hn.math_performance_index
                    if math is not None:
                        subject_scores.append(math)
                    science = edu_hn.science_performance_index
                    if science is not None:
                        subject_scores.append(science)
                    if len(subject_scores) == 0:
                        hn_score = 0
                    else:
                        hn_score = sum(subject_scores) / len(subject_scores)
                    all_scores.append(hn_score)
                df_score = pd.DataFrame(all_scores, columns=['District Scores'])
                df_score['Ranked Scores'] = df_score['District Scores'].rank(pct=True)

                # specific district nums
                subject_scores = []
                this_district = EducationCTSubgroup.objects.filter(dataset__vintage=vintages['EducationCT Subgroup'], district=district, subgroup='High Needs').first()

                if this_district is not None:
                    district_ela = this_district.ela_performance_index
                    if district_ela is not None:
                        subject_scores.append(district_ela)
                    district_math = this_district.math_performance_index
                    if district_math is not None:
                        subject_scores.append(district_math)
                    district_science = this_district.science_performance_index
                    if district_science is not None:
                        subject_scores.append(district_science)

                if len(subject_scores) == 0:
                    district_score = 'N/A'
                else:
                    district_score = sum(subject_scores) / len(subject_scores)
                base_df.at['School Performance - Econ. Disadvantaged', ('All Tracts', 'PERC')] =\
                    (df_score.loc[df_score['District Scores'] == district_score].drop_duplicates(subset=['Ranked Scores'], keep = 'first')['Ranked Scores'].to_numpy() * 100).item()

                base_df.loc['School Performance - Overall', 'Source'] = 'EDU'
                base_df.loc['School Performance - Econ. Disadvantaged', 'Source'] = 'EDU'
            except BaseException:
                print("Error creating spreadsheet for CT schools.")

        # RI
        if state_fips == '44':

            try:
                # calculate 67th percentile
                all_districts = list(EducationRI.objects.filter(dataset__vintage=vintages['EducationRI']).values('district').distinct())
                all_scores = []
                all_ED_scores = []
                for d in all_districts:
                    score = EducationRI.objects.filter(dataset__vintage=vintages['EducationRI'], district=d['district'], group='All Students').aggregate(Avg('star_rating'))['star_rating__avg']
                    ED_score = EducationRI.objects.filter(dataset__vintage=vintages['EducationRI'], district=d['district'], group='Economically Disadvantaged').aggregate(Avg('star_rating'))['star_rating__avg']
                    all_scores.append(score)
                    all_ED_scores.append(ED_score)

                df_score = pd.DataFrame(all_scores, columns=['District Scores'])
                df_ED_score = pd.DataFrame(all_ED_scores, columns=['District ED Scores'])
                df_score['Ranked Scores'] = df_score['District Scores'].rank(pct=True)
                df_ED_score['Ranked Scores'] = df_ED_score['District ED Scores'].rank(pct=True)

                # specific district nums
                district_score = EducationRI.objects.filter(dataset__vintage=vintages['EducationRI'], district=district, group='All Students').aggregate(Avg('star_rating'))['star_rating__avg']
                district_ED_score = EducationRI.objects.filter(dataset__vintage=vintages['EducationRI'], district=district, group='Economically Disadvantaged').aggregate(Avg('star_rating'))['star_rating__avg']

                base_df.at['School Performance - Overall', ('All Tracts', 'PERC')] = (df_score.loc[df_score['District Scores'] == district_score].drop_duplicates(subset = ['Ranked Scores'], keep = 'first')['Ranked Scores'].to_numpy() * 100).item()
                base_df.at['School Performance - Econ. Disadvantaged', ('All Tracts', 'PERC')] = (df_ED_score.loc[df_ED_score['District ED Scores'] == district_ED_score].drop_duplicates(subset = ['Ranked Scores'], keep = 'first')['Ranked Scores'].to_numpy() * 100).item()
                base_df.loc['School Performance - Overall', 'Source'] = 'EDU'
                base_df.loc['School Performance - Econ. Disadvantaged', 'Source'] = 'EDU'
            except BaseException:
                print("Error creating spreadsheet for RI schools.")

    def subtract_est(self, calc, first, second, base_df, col):
        assign = (col, 'EST')
        base_df.loc[calc, assign] = base_df.loc[first, assign] - base_df.loc[second, assign]
        base_df.loc[calc, 'Source'] = base_df.loc[first, 'Source'].to_numpy()

    def divide_est(self, calc, num, den, base_df, col):
        for i in range(0, len(col)):
            try:
                base_df.loc[calc, (col[i], 'EST')] = base_df.loc[num, (col[i], 'EST')] / base_df.loc[den, (col[i], 'EST')] * 100
            except ZeroDivisionError:
                base_df.loc[calc, (col[i], 'EST')] = np.nan
        base_df.loc[calc, 'Source'] = base_df.loc[num, 'Source'].to_numpy()

    def add_est(self, calc, metric_list, base_df, col):
        sums = 0
        for m in metric_list:
            sums += base_df.loc[m, (col, 'EST')]
        base_df.loc[calc, (col, 'EST')] = sums
        base_df.loc[calc, 'Source'] = base_df.loc[metric_list[0], 'Source'].to_numpy()

    def add_moe(self, calc, metric_list, base_df, col):
        if base_df.loc[metric_list[0], 'Source'].to_numpy() == 'PLACES':
            z = 1.96
        else:
            z = 1.645
        sums = 0
        for m in metric_list:
            sums += ((1/z)*base_df.loc[m, (col, 'MOE')])**2
        sums = z * sums**(1/2)
        base_df.loc[calc, (col, 'MOE')] = sums
        base_df.loc[calc, 'Source'] = base_df.loc[metric_list[0], 'Source'].to_numpy()

    def divide_moe(self, calc, num, den, frac, base_df, col):
        MOE_num = base_df.loc[num, (col, 'MOE')]
        MOE_den = base_df.loc[den, (col, 'MOE')]
        R = base_df.loc[frac, (col, 'EST')] / 100
        X_den = base_df.loc[den, (col, 'EST')]
        under_sqrt = (MOE_num**2 - (R**2).to_numpy() * MOE_den**2).to_numpy()
        if any(i <= 0 for i in X_den.to_numpy()):
            MOE_calc = np.nan
        elif all(i >= 0 for i in under_sqrt):
            MOE_calc = ((MOE_num**2 - (R**2).to_numpy() * MOE_den**2)**(1/2)).to_numpy() / X_den.to_numpy() * 100
        else:
            MOE_calc = ((MOE_num**2 + (R**2).to_numpy() * MOE_den**2)**(1/2)).to_numpy() / X_den.to_numpy() * 100

        base_df.loc[calc, (col, 'MOE')] = MOE_calc
        base_df.loc[calc, 'Source'] = base_df.loc[num, 'Source'].to_numpy()

    def agg_moe(self, tracts, base_df):
        for index, row in base_df.iterrows():
            est = base_df.loc[index, (tracts, 'EST')]
            moe = base_df.loc[index, (tracts, 'MOE')]
            est.reset_index(drop=True, inplace=True)
            moe.reset_index(drop=True, inplace=True)

            # find count of zeros
            count0 = est.isin([0]).sum()
            if count0 <= 1:
                base_df.loc[index, ('All Tracts', 'MOE')] = (((moe)**2).sum())**(1/2)
            else:
                # find location of zeros
                which0 = est.isin([0])

                # sum squared non-zeros
                non0s = ((moe[~which0])**2).sum()
                with0s = non0s + ((moe[which0])**2).max()
                moe_all = with0s**(1/2)
                base_df.loc[index, ('All Tracts', 'MOE')] = moe_all

    def row_avg(self, metric, base_df, tracts):
        sums = 0
        for t in tracts:
            sums += base_df.loc[metric, (t, 'EST')]
        base_df.loc[metric, ('All Tracts', 'EST')] = sums/len(tracts)

    def perc_avg(self, metric, base_df, tracts):
        sums = 0
        for t in tracts:
            sums += base_df.loc[metric, (t, 'PERC')]
        base_df.loc[metric, ('All Tracts', 'PERC')] = sums/len(tracts)

    def rollup_num_calc(self, metric, total_pop, base_df, tracts, state, col):
        temp = 'Total ' + metric
        sums = 0
        pop_not_nan = 0

        try:
            for t in tracts:
                if np.isnan(base_df.loc[metric, (t, 'EST')]):
                    continue
                else:
                    base_df.loc[temp, (t, 'EST')] = (base_df.loc[metric, (t, 'EST')]) * base_df.loc[total_pop, (t, 'EST')]
                    base_df.loc[temp, (t, 'MOE')] = (base_df.loc[metric, (t, 'MOE')]) * base_df.loc[total_pop, (t, 'MOE')]
                    sums += base_df.loc[temp, (t, 'EST')]
                    pop_not_nan += base_df.loc[total_pop, (t, 'EST')]
            base_df.loc[temp, ('All Tracts', 'EST')] = sums

            if pop_not_nan > 0:
                base_df.loc[metric, ('All Tracts', 'EST')] = sums / pop_not_nan
            else:
                self.log.info(f"Couldn't rollup metric={metric} because population is 0.",
                              extra={'user': self.user.username})

            # handle state
            base_df.loc[temp, (state.short_code, 'EST')] = (base_df.loc[metric, (state.short_code, 'EST')]) * \
                                                           base_df.loc[total_pop, (state.short_code, 'EST')]
            base_df.loc[temp, (state.short_code, 'MOE')] = (base_df.loc[metric, (state.short_code, 'MOE')]) * \
                                                           base_df.loc[total_pop, (state.short_code, 'MOE')]

            # aggregate
            self.agg_moe_row(tracts, [temp], base_df)
        except KeyError:
            self.log.info(f"Couldn't rollup metric={metric} because it's missing from the data.",
                          extra={'user': self.user.username})

    def rollup_percent_calc(self, metric, total_pop, base_df, tracts, state, col):
        temp = 'Total ' + metric
        sums = 0

        try:
            for t in tracts:
                base_df.loc[temp, (t, 'EST')] = (base_df.loc[metric, (t, 'EST')]) * base_df.loc[total_pop, (t, 'EST')] / 100
                base_df.loc[temp, (t, 'MOE')] = (base_df.loc[metric, (t, 'MOE')]) * base_df.loc[total_pop, (t, 'MOE')] / 100
                sums += base_df.loc[temp, (t, 'EST')]
            base_df.loc[temp, ('All Tracts', 'EST')] = sums
            base_df.loc[metric, ('All Tracts', 'EST')] = sums / base_df.loc[total_pop, ('All Tracts', 'EST')] * 100

            # handle state
            base_df.loc[temp, (state.short_code, 'EST')] = (base_df.loc[metric, (state.short_code, 'EST')]) * \
                                                           base_df.loc[total_pop, (state.short_code, 'EST')] / 100
            base_df.loc[temp, (state.short_code, 'MOE')] = (base_df.loc[metric, (state.short_code, 'MOE')]) * \
                                                           base_df.loc[total_pop, (state.short_code, 'MOE')] / 100

            # aggregate
            self.agg_moe_row(tracts, [temp], base_df)
        except BaseException as e:
            print("Error rolling up percent calc")

    def divide_rows(self, calc, num, den, frac, base_df, col):
        self.divide_est(calc, num, den, base_df, col)
        self.divide_moe_all_only(calc, num, den, frac, base_df)

    def divide_est(self, calc, num, den, base_df, col):
        for i in range(0, len(col)):
            try:
                base_df.loc[calc, (col[i], 'EST')] = base_df.loc[num, (col[i], 'EST')] / base_df.loc[den, (col[i], 'EST')] * 100
            except ZeroDivisionError:
                base_df.loc[calc, (col[i], 'EST')] = np.nan
        base_df.loc[calc, 'Source'] = base_df.loc[num, 'Source'].to_numpy()

    def divide_moe_all_only(self, calc, num, den, frac, base_df):
        MOE_num = base_df.loc[num, ('All Tracts', 'MOE')]
        MOE_den = base_df.loc[den, ('All Tracts', 'MOE')]
        R = base_df.loc[frac, ('All Tracts', 'EST')] / 100
        X_den = base_df.loc[den, ('All Tracts', 'EST')]
        under_sqrt = (MOE_num**2 - (R**2) * MOE_den**2)
        if under_sqrt >= 0:
            MOE_calc = ((MOE_num**2 - (R**2) * MOE_den**2)**(1/2)) / X_den* 100
        else:
            MOE_calc = ((MOE_num**2 + (R**2) * MOE_den**2)**(1/2)) / X_den * 100
        base_df.loc[calc, ('All Tracts', 'MOE')] = MOE_calc
        base_df.loc[calc, 'Source'] = base_df.loc[num, 'Source'].to_numpy()

    def agg_moe_row(self, tracts, metric_list, base_df):
        for m in metric_list:
            est = base_df.loc[m, (tracts, 'EST')]
            moe = base_df.loc[m, (tracts, 'MOE')]
            est.reset_index(drop=True, inplace=True)
            moe.reset_index(drop=True, inplace=True)
            #find count of zeros
            count0 = est.isin([0]).sum()
            if count0 <= 1:
                base_df.loc[m, ('All Tracts', 'MOE')] = (((moe)**2).sum())**(1/2)
            else:
                # find location of zeros
                which0 = est.isin([0])

                # sum squared non-zeros
                non0s = ((moe[~which0])**2).sum()
                with0s = non0s + ((moe[which0])**2).max()
                moe_all = with0s**(1/2)
                base_df.loc[m, ('All Tracts', 'MOE')] = moe_all


class EST5EDT(tzinfo):

    def utcoffset(self, dt):
        return timedelta(hours=-5) + self.dst(dt)

    def dst(self, dt):
        d = datetime(dt.year, 3, 8)        #2nd Sunday in March
        self.dston = d + timedelta(days=6-d.weekday())
        d = datetime(dt.year, 11, 1)       #1st Sunday in Nov
        self.dstoff = d + timedelta(days=6-d.weekday())
        if self.dston <= dt.replace(tzinfo=None) < self.dstoff:
            return timedelta(hours=1)
        else:
            return timedelta(0)

    def tzname(self, dt):
        return 'EST5EDT'
